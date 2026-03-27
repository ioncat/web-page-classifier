import sqlite3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

DB_PATH = r"E:\My files\0 My_Dev\url-parser\urls.db"
OUT_PATH = r"E:\My files\0 My_Dev\url-parser\taxonomy_review.xlsx"

FONT = Font(name="Arial", size=10)
FONT_BOLD = Font(name="Arial", size=10, bold=True)

FILL_RED    = PatternFill(patternType="solid", fgColor="FFFFB3B3")
FILL_YELLOW = PatternFill(patternType="solid", fgColor="FFFFFF99")
FILL_WHITE  = PatternFill(patternType="solid", fgColor="FFFFFFFF")

def set_row_font(row_cells, font):
    for cell in row_cells:
        cell.font = font

def make_sheet1(wb, conn):
    ws = wb.create_sheet("Категории")
    headers = ["Категория", "Кол-во", "Пример URL"]
    ws.append(headers)

    # Bold header
    for cell in ws[1]:
        cell.font = FONT_BOLD
        cell.alignment = Alignment(wrap_text=False)

    ws.freeze_panes = "A2"

    sql = """
        SELECT category, COUNT(*) as cnt, MIN(url) as example_url
        FROM urls
        WHERE category IS NOT NULL
        GROUP BY category
        ORDER BY cnt DESC
    """
    cur = conn.cursor()
    cur.execute(sql)
    rows = cur.fetchall()

    for category, cnt, example_url in rows:
        ws.append([category, cnt, example_url])
        row_idx = ws.max_row

        # Determine fill
        cat = category or ""
        is_red = (
            "URL:" in cat
            or "Category:" in cat
            or (len(cat) > 0 and cat[0].isdigit())
            or len(cat) > 60
        )
        is_yellow = (cnt == 1)

        if is_red:
            fill = FILL_RED
        elif is_yellow:
            fill = FILL_YELLOW
        else:
            fill = FILL_WHITE

        for cell in ws[row_idx]:
            cell.font = FONT
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=False)

    # Column widths
    ws.column_dimensions[get_column_letter(1)].width = 50
    ws.column_dimensions[get_column_letter(2)].width = 10
    ws.column_dimensions[get_column_letter(3)].width = 60

    return ws


def make_sheet2(wb, conn):
    ws = wb.create_sheet("URL")
    headers = ["Категория", "Заголовок", "URL"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = FONT_BOLD
        cell.alignment = Alignment(wrap_text=False)

    ws.freeze_panes = "A2"

    sql = """
        SELECT category, title, url
        FROM urls
        WHERE category IS NOT NULL
        ORDER BY category, title
    """
    cur = conn.cursor()
    cur.execute(sql)
    rows = cur.fetchall()

    for category, title, url in rows:
        ws.append([category, title, url])
        row_idx = ws.max_row
        for cell in ws[row_idx]:
            cell.font = FONT
            cell.alignment = Alignment(wrap_text=False)

    # Column widths
    ws.column_dimensions[get_column_letter(1)].width = 35
    ws.column_dimensions[get_column_letter(2)].width = 50
    ws.column_dimensions[get_column_letter(3)].width = 70

    return ws


def main():
    conn = sqlite3.connect(DB_PATH)
    wb = openpyxl.Workbook()

    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    make_sheet1(wb, conn)
    make_sheet2(wb, conn)

    conn.close()
    wb.save(OUT_PATH)
    print("Done: taxonomy_review.xlsx")


if __name__ == "__main__":
    main()
